#!/usr/bin/env python3
"""Generate a DeskPricer demo Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "DeskPricer_Bitcoin_Demo.xlsx"

# ------------------------------------------------------------------
# Styles
# ------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LABEL_FONT = Font(bold=True)
VALUE_FONT = Font()
NOTE_FONT = Font(italic=True, color="666666", size=9)
FORMULA_FONT = Font(color="0563C1")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_range(ws, start_row, start_col, end_row, end_col, font=None, fill=None, border=None, alignment=None):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def _add_header(ws, row, col, text, width=20):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width


def _add_label_value(ws, row, label, value, col=1):
    ws.cell(row=row, column=col, value=label).font = LABEL_FONT
    ws.cell(row=row, column=col + 1, value=value).font = VALUE_FONT


def _add_output_row(ws, row, label, xpath, expected, url_cell="$B$14"):
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    formula = f'=VALUE(FILTERXML({url_cell},"//outputs/{xpath}"))'
    ws.cell(row=row, column=2, value=formula).font = FORMULA_FONT
    ws.cell(row=row, column=3, value=expected).font = NOTE_FONT
    ws.cell(row=row, column=4, value="Expected").font = NOTE_FONT


# ------------------------------------------------------------------
# Sheet 1: Greeks
# ------------------------------------------------------------------
def build_greeks_sheet(wb):
    ws = wb.active
    ws.title = "Greeks"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "DeskPricer Demo — Bitcoin European Call (3M)"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Inputs section
    ws.merge_cells("A3:D3")
    ws["A3"] = "Inputs"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    inputs = [
        ("Spot (s)", 75000),
        ("Strike (k)", 100000),
        ("Time to expiry (t)", 0.25),
        ("Risk-free rate (r)", 0.0365),
        ("Dividend yield (q)", 0),
        ("Volatility (v)", 0.5),
        ("Type", "call"),
        ("Style", "european"),
    ]
    for i, (label, val) in enumerate(inputs, start=4):
        _add_label_value(ws, i, label, val)

    # URL construction
    row = 13
    ws.cell(row=row, column=1, value="Service URL").font = LABEL_FONT
    url_formula = (
        '="http://127.0.0.1:8765/v1/greeks?s="&B4&"&k="&B5&"&t="&B6&'
        '"&r="&B7&"&q="&B8&"&v="&B9&"&type="&B10&"&style="&B11'
    )
    ws.cell(row=row, column=2, value=url_formula).font = FORMULA_FONT

    row = 14
    ws.cell(row=row, column=1, value="Raw XML").font = LABEL_FONT
    ws.cell(row=row, column=2, value="=WEBSERVICE(B13)").font = FORMULA_FONT

    # Outputs section
    row = 16
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Outputs").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL

    outputs = [
        ("Price", "price", 1423.850543081),
        ("Delta", "delta", 0.160835646),
        ("Gamma", "gamma", 1.3039e-05),
        ("Vega", "vega", 91.427494809),
        ("Theta", "theta", -26.181325921),
        ("Rho", "rho", 26.524188656),
        ("Charm", "charm", -0.001769148),
    ]
    for i, (label, xpath, expected) in enumerate(outputs, start=17):
        _add_output_row(ws, i, label, xpath, expected, url_cell="$B$14")

    # Note
    row = 25
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Note: Ensure DeskPricer is running on 127.0.0.1:8765 before opening this file.").font = NOTE_FONT


# ------------------------------------------------------------------
# Sheet 2: Implied Vol
# ------------------------------------------------------------------
def build_impliedvol_sheet(wb):
    ws = wb.create_sheet(title="ImpliedVol")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    ws.merge_cells("A1:D1")
    ws["A1"] = "DeskPricer Demo — Implied Volatility Solver"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A3:D3")
    ws["A3"] = "Inputs"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL

    inputs = [
        ("Spot (s)", 75000),
        ("Strike (k)", 100000),
        ("Time to expiry (t)", 0.25),
        ("Risk-free rate (r)", 0.0365),
        ("Dividend yield (q)", 0),
        ("Market Price", 3398.71),
        ("Type", "call"),
        ("Style", "european"),
    ]
    for i, (label, val) in enumerate(inputs, start=4):
        _add_label_value(ws, i, label, val)

    row = 13
    ws.cell(row=row, column=1, value="Service URL").font = LABEL_FONT
    url_formula = (
        '="http://127.0.0.1:8765/v1/impliedvol?s="&B4&"&k="&B5&"&t="&B6&'
        '"&r="&B7&"&q="&B8&"&price="&B9&"&type="&B10&"&style="&B11'
    )
    ws.cell(row=row, column=2, value=url_formula).font = FORMULA_FONT

    row = 14
    ws.cell(row=row, column=1, value="Raw XML").font = LABEL_FONT
    ws.cell(row=row, column=2, value="=WEBSERVICE(B13)").font = FORMULA_FONT

    row = 16
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Outputs").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL

    _add_output_row(ws, 17, "Implied Vol", "implied_vol", 0.682835198, url_cell="$B$14")
    _add_output_row(ws, 18, "NPV @ IV", "npv_at_iv", 3398.710167331, url_cell="$B$14")

    row = 20
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Note: Market price of 3398.71 implies ~68.3% annualized vol.").font = NOTE_FONT


# ------------------------------------------------------------------
# Sheet 3: PnL Attribution
# ------------------------------------------------------------------
def build_pnl_sheet(wb):
    ws = wb.create_sheet(title="PnL Attribution")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    ws.merge_cells("A1:E1")
    ws["A1"] = "DeskPricer Demo — PnL Attribution (t-1 → t)"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # t-1 inputs
    ws.merge_cells("A3:C3")
    ws["A3"] = "t-1 Inputs"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL
    t1_inputs = [
        ("Spot s_t_minus_1", 75000),
        ("Time t_t_minus_1", 0.25),
        ("Rate r_t_minus_1", 0.0365),
        ("Div q_t_minus_1", 0),
        ("Vol v_t_minus_1", 0.5),
    ]
    for i, (label, val) in enumerate(t1_inputs, start=4):
        _add_label_value(ws, i, label, val)

    # t inputs
    ws.merge_cells("D3:E3")
    ws["D3"] = "t Inputs"
    ws["D3"].font = SUBHEADER_FONT
    ws["D3"].fill = SUBHEADER_FILL
    t_inputs = [
        ("Spot s_t", 80000),
        ("Time t_t", 0.2466),
        ("Rate r_t", 0.0365),
        ("Div q_t", 0),
        ("Vol v_t", 0.55),
    ]
    for i, (label, val) in enumerate(t_inputs, start=4):
        ws.cell(row=i, column=4, value=label).font = LABEL_FONT
        ws.cell(row=i, column=5, value=val).font = VALUE_FONT

    # Shared inputs
    row = 10
    ws.cell(row=row, column=1, value="Strike (k)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=100000).font = VALUE_FONT
    ws.cell(row=row, column=4, value="Quantity").font = LABEL_FONT
    ws.cell(row=row, column=5, value=1).font = VALUE_FONT
    row = 11
    ws.cell(row=row, column=1, value="Type").font = LABEL_FONT
    ws.cell(row=row, column=2, value="call").font = VALUE_FONT
    ws.cell(row=row, column=4, value="Style").font = LABEL_FONT
    ws.cell(row=row, column=5, value="european").font = VALUE_FONT

    # URL
    row = 13
    ws.cell(row=row, column=1, value="Service URL").font = LABEL_FONT
    url_formula = (
        '="http://127.0.0.1:8765/v1/pnl_attribution?s_t_minus_1="&B4&"&s_t="&D4&'
        '"&k="&B10&"&t_t_minus_1="&B5&"&t_t="&D5&"&r_t_minus_1="&B6&"&r_t="&D6&'
        '"&q_t_minus_1="&B7&"&q_t="&D7&"&v_t_minus_1="&B8&"&v_t="&D8&'
        '"&type="&B11&"&style="&E11&"&qty="&E10'
    )
    ws.cell(row=row, column=2, value=url_formula).font = FORMULA_FONT

    row = 14
    ws.cell(row=row, column=1, value="Raw XML").font = LABEL_FONT
    ws.cell(row=row, column=2, value="=WEBSERVICE(B13)").font = FORMULA_FONT

    # Outputs
    row = 16
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="Outputs").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL

    outputs = [
        ("Price t-1", "price_t_minus_1", 1423.850543081),
        ("Price t", "price_t", 2989.673408685),
        ("Actual PnL", "actual_pnl", 1565.822865604),
        ("Delta PnL", "delta_pnl", 804.178231041),
        ("Gamma PnL", "gamma_pnl", 162.98430088),
        ("Vega PnL", "vega_pnl", 457.137474045),
        ("Theta PnL", "theta_pnl", -26.181325921),
        ("Rho PnL", "rho_pnl", 0),
        ("Vanna PnL", "vanna_pnl", 0),
        ("Volga PnL", "volga_pnl", 0),
        ("Explained PnL", "explained_pnl", 1398.118680045),
        ("Residual PnL", "residual_pnl", 167.704185559),
    ]
    for i, (label, xpath, expected) in enumerate(outputs, start=17):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        formula = f'=VALUE(FILTERXML($B$14,"//outputs/{xpath}"))'
        ws.cell(row=i, column=2, value=formula).font = FORMULA_FONT
        ws.cell(row=i, column=3, value=expected).font = NOTE_FONT
        ws.cell(row=i, column=4, value="Expected").font = NOTE_FONT

    row = 30
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="Note: Vanna/Volga are zero because cross_greeks=false by default. Add &cross_greeks=true to the URL to enable.").font = NOTE_FONT


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def main():
    wb = Workbook()
    build_greeks_sheet(wb)
    build_impliedvol_sheet(wb)
    build_pnl_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
